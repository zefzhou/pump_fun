from pump_fun import PumpFun
from config import *
import concurrent


def split_array(array, group_size):
    start = 0
    end = len(array)
    groups = []
    while start + group_size < end:
        groups.append(array[start:start + group_size])
        start += group_size
    if start < end:
        groups.append(array[start:end])
    return groups


def which_tokens(shuffle, start, batch_size, total_size):
    start += shuffle * batch_size
    if start < total_size - batch_size:
        return list(range(start, start + batch_size))
    else:
        return list(range(start, total_size)) + list(
            range(0, start + batch_size - total_size))


def get_pk_list(file_path: str = 'accounts.txt'):
    pk_list = []
    with open(file_path, 'r') as f:
        lines = f.readlines()
        for line in lines:
            pk = line.strip()
            pk_list.append(pk)
    return pk_list


def get_tokens(file_path: str = 'tokens.xlsx'):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    token_list = []
    for row in data[1:]:
        token = row[1]
        if str(token).startswith('http'):
            token = str(token).split('https://pump.fun/')[1]
        token_list.append(token)
    return token_list


def trade(pk: str, tokens: list):
    pf = PumpFun(private_key=pk)
    for token in tokens:
        try:
            pf.trade(token_addr=token)
        except Exception as e:
            print(e)


def main():
    pk_list = get_pk_list()
    token_list = get_tokens()
    tokens_list = split_array(token_list, TOKENS_PER_WALLET)

    with concurrent.futures.ThreadPoolExecutor(
            max_workers=CONCURRENT) as executor:
        futures = [
            executor.submit(trade, pk_list[i],
                            tokens_list[(SHUFFLE + i) % len(tokens_list)])
            for i in range(len(pk_list))
        ]
        concurrent.futures.wait(futures)


if __name__ == '__main__':
    main()
